import pandas as pd
from openpyxl import load_workbook

# File path to your Excel file
file_path = r'C:\Users\pr38\Downloads\SNOWFLAKE_STTM_LSAMS (1).xlsx'

# Load workbook using openpyxl
wb = load_workbook(file_path, data_only=True)
ws = wb['Specific table']  # Specify the correct sheet name

# Load the Excel sheet into a DataFrame
df = pd.read_excel(file_path, sheet_name='Specific table')

# Mapping for table names
has = {
    'SRVEOMX': 'LOAN_EOM_SNAPSHOT',
    'SRVCHGI': 'INVESTOR_PENDING_TRANSFER',
    'SRVCPLAN': 'CORPORATE_BILLING_PLAN',
    'LSBKAOBD00': 'Bankruptcy_Paid_Order_Detail',
    'SRVFBDUE': 'LOAN_LATE_FEE',
    'IVMGT00': 'Invoice_Management',
    'DMDRDH00': 'Demand_Letter',
    'LSBKCAS00': 'Bankruptcy_Case_Track',
    'LSCLLN00': 'LOAN_MILESTONE_DATE',
    'LSSTMX20': 'LOAN_MONTH_STATEMENT_TRANSMISSION',
    'SRVCHGV': 'LOAN_ARM',
    'SRVESCE': 'LOAN_ESCROW',
    'SRVCOLI': 'LOAN_COLLECTION_ITEM',
    'LSFBDTL00': 'FOREBEARANCE_PAYMENT_DETAIL'
}

table_scripts = {}

grouped = df.groupby('LSAMS TABLE ')

for table_name, group in grouped:
    tbl_nm = has.get(table_name, table_name)
    create_stmt = f"CREATE TABLE {table_name} (\n"
    columns = []
    primary_keys = []

    for index, row in group.iterrows():
        column_name = row['LSAMS COLUMN ']
        datatype = row['DATA TYPE ']


        cell = ws.cell(row=index + 2, column=group.columns.get_loc('LSAMS COLUMN ') + 1)


      
        fill_color = cell.fill.start_color.rgb
        if fill_color == 'FFFF0000':  
            print(f"Primary key detected: {column_name}")
            primary_keys.append(column_name)
        
        columns.append(f"    {column_name} {datatype}")

    create_stmt += ",\n".join(columns)
    
    if primary_keys:
        create_stmt += f",\n CONSTRAINT PK_{table_name} PRIMARY KEY ({', '.join(primary_keys)})"
    
    create_stmt += "\n);"
    table_scripts[tbl_nm] = create_stmt

output_file = r'C:\Users\pr38\Downloads\lsams_tbl_list.txt'
for table_name, script in table_scripts.items():
    with open(output_file, 'a') as file:
        file.write(script + '\n')
